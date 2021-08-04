import matplotlib.pyplot as plt
import pandas as pd
import os
import re
from openpyxl import load_workbook
from tqdm import tqdm
import numpy as np
from itertools import chain
from scipy.optimize import curve_fit
import math
# import PyGnuplot


def show_lineage_trajectories(file_name):
    sim = pd.read_csv(file_name)
    for index, row in sim.iterrows():
        row = row.to_numpy()
        plt.plot(row)
    plt.ylabel('density')
    plt.xlabel('generation')
    plt.show()


def AF_test(file_name, gen_num_to_average):
    df = pd.read_csv(file_name)
    densities = []
    for generation, column in df.iteritems():
        generation_mean = column.mean()
        densities.append(generation_mean)
    final_mean = sum(densities[len(densities) - gen_num_to_average:len(densities)]) / gen_num_to_average
    return final_mean


def show_kymograph(file_name, out_put_folder):
    df = pd.read_csv(file_name)
    x = []
    y = []
    for index, row in df.iterrows():
        row = row.to_numpy()
        if row.sum() != 0:
            x_coordinates = np.argwhere(row == 1).flatten().tolist()
            x.append(x_coordinates)
            # y_coordinates = [index * 20] * row.sum()
            y_coordinates = [index] * row.sum()
            y.append(y_coordinates)
        else:
            break
    x = list(chain.from_iterable(x))
    y = list(chain.from_iterable(y))
    plt.hist2d(x, y, bins=(len(row), index+1), cmap='gray')
    plt.xlabel('nucleosome position')
    plt.ylabel('generation')
    output_file_name = 'D_init=' + re.split('\.|_', file_name)[-3]
    plt.savefig(out_put_folder + '/' + output_file_name + '.png', dpi=200)
    plt.clf()


def get_parameters(path_name):
    dir_ls = os.listdir(path_name)
    Ss = []
    RRs = []
    AFs = []
    CFs = []
    for file in dir_ls:
        S = re.split('\.|_', file)[0][0:2]
        while len(S) < 2:
            S = S + '0'
        S = float(S) / 10
        Ss.append(S)
        RR = re.split('\.|_', file)[1]
        RR = int(RR)
        RRs.append(RR)
        AF = re.split('\.|_', file)[2][0:5]
        while len(AF) < 5:
            AF = AF + '0'
        AF = float(AF) / 10000
        AF = round(AF, 3)
        AFs.append(AF)
        CF = re.split('\.|_', file)[3][0:5]
        while len(CF) < 5:
            CF = CF + '0'
        CF = float(CF) / 10000
        CF = round(CF, 3)
        CFs.append(CF)
    return Ss, RRs, AFs, CFs


def get_file_names(path_name):
    dir_ls = os.listdir(path_name)
    file_names = []
    for file in dir_ls:
        file_name = path_name + file
        file_names.append(file_name)
    return file_names


def save_AF_test_excel(path_name, excel_name, sheet_name):
    book = load_workbook(excel_name)
    writer = pd.ExcelWriter(excel_name, engine='openpyxl')
    writer.book = book
    file_names = get_file_names(path_name)
    Ss, RRs, AFs, CFs = get_parameters(path_name)
    x = AFs
    y = []
    for i in tqdm(range(len(x))):
        file_name = file_names[i]
        res = AF_test(file_name, 2000)
        y.append(res)
    df = pd.DataFrame([x, y]).transpose()
    df.to_excel(writer, sheet_name=sheet_name)
    writer.save()
    writer.close()


def output_kymographs(path_name, output_folder):
    file_names = get_file_names(path_name)
    for file_name in file_names:
        if file_name[-5] == '0':
            show_kymograph(file_name, output_folder)


def get_exp_fit_parameters(file_name):
    df = pd.read_csv(file_name)
    densities = []
    for generation, column in df.iteritems():
        generation_mean = column.mean()
        densities.append(generation_mean)
    gen_num_to_average = 500
    final_mean = sum(densities[len(densities) - gen_num_to_average:len(densities)]) / gen_num_to_average
    densities -= final_mean
    x = np.linspace(0, 29, 30)
    y = np.array(densities)
    y -= final_mean
    y[y < 0] *= -1
    y = y[0:30]
    fit = curve_fit(lambda t, a, b: a * np.exp(b * t), x, y, p0=(0, 0))
    a = fit[0][0]
    b = fit[0][1]
    return a, b, final_mean


# path_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/replenish_efficiency/af=0.7/raw_data/replenish_efficiency/'
# Ss, RRs, AFs, CFs = get_parameters(path_name)
# file_names = get_file_names(path_name)
# zipped_ls = sorted(zip(CFs, file_names))
# cm = plt.get_cmap('gist_rainbow')
# counter = 0
# for i in zipped_ls:
#     CF = i[0]
#     file_name = i[1]
#     df = pd.read_csv(file_name)
#     re_means = []
#     for generation, column in df.iteritems():
#         re_mean = column.mean()
#         re_means.append(re_mean)
#     plt.plot(re_means, label='CF={}'.format(CF), color=cm(1.*counter/len(CFs)))
#     counter += 1
# # plt.xscale('log')
# plt.xlim((-1, 20))
# plt.yscale('log')
# plt.title('S={} RR={} AF={}'.format(Ss[0], RRs[0], AFs[0]))
# plt.xlabel('generation')
# plt.ylabel('replenish efficiency')
# plt.legend(loc='upper right')
# plt.show()


# excel_name = '/Users/kikawaryoku/Desktop/100_3_02-05_AF_test.xlsx'
# df = pd.read_excel(excel_name, sheet_name=0)
# plt.plot(df.AF, df.density, '.', label='10,000')
# df = pd.read_excel(excel_name, sheet_name=1)
# plt.plot(df.AF, df.density, '.', label='100,000')
# plt.title('S=100 RR=3')
# plt.ylabel('density')
# plt.xlabel('loading efficiency')
# plt.legend()
# plt.show()
# save_AF_test_excel(path_name, excel_name, '100000')
# path_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/local_incorp_res/3_10_014-015_100000/raw_data/nuc_comp/'
# output_folder = '/Users/kikawaryoku/Desktop/3_10_kymographs'
# output_kymographs(path_name, output_folder)

# path_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/replenish_efficiency/af=0.7_reversed/raw_data/density/'
# excel_name = '/Users/kikawaryoku/Desktop/reversed.xlsx'
# book = load_workbook(excel_name)
# writer = pd.ExcelWriter(excel_name, engine='openpyxl')
# writer.book = book
# file_names = get_file_names(path_name)
# Ss, RRs, AFs, CFs = get_parameters(path_name)
# a_ls = []
# b_ls = []
# D0_ls = []
# for file_name in file_names:
#     a, b, D0 = get_exp_fit_parameters(file_name)
#     a_ls.append(a)
#     b_ls.append(b)
#     D0_ls.append(D0)
# CFs = np.array(CFs)
# D0_a = np.array(D0_ls)
# deviations = CFs - D0_ls
# deviations = list(deviations)
# df = pd.DataFrame([deviations, a_ls, b_ls]).transpose()
# df.to_excel(writer, sheet_name='AF=0.7')
# writer.save()
# writer.close()

# plt.plot(deviations, a_ls, '.', label='AF=0.55')

# path_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/replenish_efficiency/af=0.7/raw_data/density/'
# file_names = get_file_names(path_name)
# Ss, RRs, AFs, CFs = get_parameters(path_name)
# a_ls = []
# D0_ls = []
# for file_name in file_names:
#     a, D0 = get_a(file_name)
#     a_ls.append(a)
#     D0_ls.append(D0)
# CFs = np.array(CFs)
# D0_a = np.array(D0_ls)
# deviations = CFs - D0_ls
# plt.plot(deviations, a_ls, '.', label='AF=0.7')
#
# path_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/replenish_efficiency/af=0.55/raw_data/density/'
# file_names = get_file_names(path_name)
# Ss, RRs, AFs, CFs = get_parameters(path_name)
# a_ls = []
# D0_ls = []
# for file_name in file_names:
#     a, b, D0 = get_exp_fit_parameters(file_name)
#     a_ls.append(a)
#     D0_ls.append(D0)
# # CFs = np.array(CFs)
# D0_a = np.array(D0_ls)
# print(D0_a.mean())
# deviations = CFs - D0_ls
# plt.plot(deviations, a_ls, '.', label='AF=1')
#
# plt.title('D - D0 = A * e^(B * t)')
# plt.ylabel('A')
# plt.xlabel('D_init - D0')
# plt.legend()
# plt.show()

# excel_name = '/Users/kikawaryoku/Desktop/detailed.xlsx'
# df = pd.read_excel(excel_name, sheet_name=3)
# x = np.linspace(0, 30, 31)
# for index, row in df.iterrows():
#     if row.deviation > 0:
#         y = row.A * np.exp(row.B * x)
#     else:
#         y = -1 * row.A * np.exp(row.B * x)
#     y = y + 0.225
#     plt.plot(x, y)
# plt.xlabel('generation')
# plt.ylabel('density')
# plt.title('AF=0.7 (exponetial fits)')
# plt.show()

# path_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/replenish_efficiency/af=0.7_reversed/raw_data/density/'
# file_names = get_file_names(path_name)
# Ss, RRs, AFs, CFs = get_parameters(path_name)
# for file_name in file_names:
#     df = pd.read_csv(file_name)
#     densities = []
#     for generation, column in df.iteritems():
#         generation_mean = column.mean()
#         densities.append(generation_mean)
#     plt.plot(densities[0:31])
# plt.title('AF=0.7 (1st dilute 2nd replenish)')
# plt.xlabel('generation')
# plt.ylabel('density')
# plt.show()

# excel_name = '/Users/kikawaryoku/Desktop/detailed.xlsx'
# df = pd.read_excel(excel_name, sheet_name=1)
# plt.plot(df.deviation, df.B, '.', label='AF=0.7')
# df = pd.read_excel(excel_name, sheet_name=3)
# plt.plot(df.deviation, df.B, '.', label='AF=0.7 long')
# plt.title('D - D0 = A * e^(B * t)')
# plt.xlabel('D_init - D0')
# plt.ylabel('B')
# plt.legend()
# plt.show()

# path_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/kymographs_vs_CFs/AF=0.5/raw_data/nuc_comp/'
# output_folder = '/Users/kikawaryoku/Desktop/AF=0.5'
# output_kymographs(path_name, output_folder)

# excel_name = '/Users/kikawaryoku/OneDrive - University of Edinburgh/CENP-A/analyzed_results/07_05_2020/detailed.xlsx'
# df = pd.read_excel(excel_name, sheet_name=1)
# cm = plt.get_cmap('gist_rainbow')
# for i, row in df.iterrows():
#     # if row.deviation > 0:
#     #     a = row.A
#     # else:
#     #     a = - row.A
#     if row.deviation < -0.004:
#         a = row.A
#         b = row.B
#         dev = row.deviation
#         x = np.linspace(0, 30)
#         y = a * math.e ** (b * x)
#         plt.plot(x, y)
# plt.semilogy()
# plt.title('D_init < D0')
# plt.ylabel('|D - D0|')
# plt.ylim(0.004, 0.11)
# plt.xlim(-1, 10)
# plt.xlabel('generation')
# plt.show()

# path_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/replenish_efficiency/af=0.7_detailed/raw_data/density/'
# file_names = get_file_names(path_name)
# Ss, RRs, AFs, CFs = get_parameters(path_name)
# D0_ls = []
# counter = 0
# for file_name in file_names:
#     a, b, D0 = get_exp_fit_parameters(file_name)
#     D0_ls.append(D0)
#     D0_a = np.array(D0_ls)
#     df = pd.read_csv(file_name)
#     densities = []
#     for generation, column in df.iteritems():
#         generation_mean = column.mean()
#         densities.append(generation_mean)
#     subtracted_densities = densities[0:31]
#     subtracted_densities = np.array(subtracted_densities)
#     subtracted_densities = subtracted_densities - D0_a.mean()
#     if subtracted_densities[0] > 0:
#         for i in range(len(subtracted_densities)):
#             if subtracted_densities[i] < 0:
#                 subtracted_densities[i] = -1 * subtracted_densities[i]
#         plt.plot(subtracted_densities)
#     counter += 1
# plt.title('D_init > D0')
# plt.xlabel('generation')
# plt.ylabel('|D - D0|')
# plt.semilogy()
# plt.xlim(-1, 10)
# plt.ylim(0.004, 0.11)
# plt.show()

# path_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/kymographs_vs_CFs/AF=0.7/raw_data/nuc_comp/'
# output_folder = '/Users/kikawaryoku/Desktop/kymographs'
# output_kymographs(path_name, output_folder)


# excel_name = '/Users/kikawaryoku/OneDrive - University of Edinburgh/CENP-A/analyzed_results/07_05_2020/detailed.xlsx'
# df = pd.read_excel(excel_name, sheet_name=1)
# plt.scatter(df.deviation, df.B, marker='.')
# excel_name = '/Users/kikawaryoku/OneDrive - University of Edinburgh/CENP-A/analyzed_results/07_05_2020/broad.xlsx'
# df_1 = pd.read_excel(excel_name, sheet_name=3)
# plt.scatter(df_1.deviation, df_1.B, marker='.')
# plt.xlabel('D_init - D0')
# plt.ylabel('B')
# plt.ylim(-1.6, 0)
# plt.title('AF=0.7 manual')
# plt.show()

# path_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/replenish_efficiency/af=0.7/raw_data/density/'
# excel_name = '/Users/kikawaryoku/OneDrive - University of Edinburgh/CENP-A/analyzed_results/07_05_2020/broad.xlsx'
# df_1 = pd.read_excel(excel_name, sheet_name=3)
# file_names = get_file_names(path_name)
# Ss, RRs, AFs, CFs = get_parameters(path_name)
# D0_ls = []
# counter = 0
# for file_name in file_names:
#     a, b, D0 = get_exp_fit_parameters(file_name)
#     D0_ls.append(D0)
#     D0_a = np.array(D0_ls)
#     df = pd.read_csv(file_name)
#     densities = []
#     for generation, column in df.iteritems():
#         generation_mean = column.mean()
#         densities.append(generation_mean)
#     subtracted_densities = densities[0:31]
#     subtracted_densities = np.array(subtracted_densities)
#     subtracted_densities = subtracted_densities - D0_a.mean()
#     for i in range(len(subtracted_densities)):
#         if subtracted_densities[i] < 0:
#             subtracted_densities[i] = -1 * subtracted_densities[i]
#     plt.plot(subtracted_densities, '.', label='observed')
#     for index, row in df_1.iterrows():
#         if CFs[counter] - 0.01 < row.deviation + D0_a.mean() < CFs[counter] + 0.01:
#             a = row.A
#             b = row.B
#             dev = row.deviation
#             x = np.linspace(0, 30)
#             y = a * math.e ** (b * x)
#             plt.plot(x, y, label='predicted')
#     plt.title('Initial density={}'.format(CFs[counter]))
#     plt.semilogy()
#     # plt.xlim(-0.5, 30.5)
#     plt.ylim(0.0001, 1)
#     plt.xlabel('generation')
#     plt.ylabel('|D - D0|')
#     plt.legend()
#     destination_file_name = '/Users/kikawaryoku/Desktop/23_06_2020/manual_full/{}.png'.format(CFs[counter])
#     plt.savefig(destination_file_name, dpi=200)
#     plt.clf()
#     counter += 1

file_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/length&stochasticity/af=0.7_nn=200/raw_data/density/3_3_07_002.csv'
df = pd.read_csv(file_name)
means = []
for index, column in df.iteritems():
    column_mean = column.mean()
    means.append(column_mean)
plt.plot(means, label='NN=200')

file_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/length&stochasticity/af=0.7_nn=500/raw_data/density/3_3_07_002.csv'
df = pd.read_csv(file_name)
means = []
for index, column in df.iteritems():
    column_mean = column.mean()
    means.append(column_mean)
plt.plot(means, label='NN=500')

file_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/length&stochasticity/af=0.7_nn=2500/raw_data/density/3_3_07_002.csv'
df = pd.read_csv(file_name)
means = []
for index, column in df.iteritems():
    column_mean = column.mean()
    means.append(column_mean)
plt.plot(means, label='NN=2500')

file_name = '/Users/kikawaryoku/PycharmProjects/project_cenp_a/length&stochasticity/af=0.7_nn=5000/raw_data/density/3_3_07_002.csv'
df = pd.read_csv(file_name)
means = []
for index, column in df.iteritems():
    column_mean = column.mean()
    means.append(column_mean)
plt.plot(means, label='NN=5000')
plt.xlabel('generation')
plt.ylabel('density')
plt.ylim(0, 1)
plt.xlim(-50, 500)
plt.legend()
plt.show()
